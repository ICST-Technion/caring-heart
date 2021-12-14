from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string as _col_idx
import operator
from extensions.MSingleton import Singleton
from extensions.MCExtensions import with_metaclass
from inventory.db.sheetname_getter import sheetname_getter
from inventory.interfaces.IDb import IDb
from itertools import islice
from configparser import ConfigParser


def col_idx(letter):
    return _col_idx(letter) - 1


def skip_first(iterable):
    return islice(iterable, 1, None)


# class ExcelDb(IDb, metaclass=with_metaclass(IDb.__class__, Singleton)):
class ExcelDb(IDb):

    def __init__(self,
                 *args,
                 filename="template.xlsx",
                 sheetname_getter=sheetname_getter,
                 key_column_map=None,
                 workbook_loader=None,
                 boolean_keys=(),
                 **kwargs):
        if key_column_map is None:
            key_column_map = {"name": 'A', "address": 'B', "street": 'C', "city": 'D', "phone": 'E', "description": 'F',
                       "category": 'G', "date": 'H', "comments": 'I', "email": 'J'}
        if workbook_loader is None:
            workbook_loader = lambda: load_workbook(filename=self._filename)

        assert all(map(lambda col: col in key_column_map.keys(), boolean_keys))
        self._boolean_keys = boolean_keys

        self._workbook_loader = workbook_loader
        self._filename = filename
        self._wb = None
        self._sht = None
        self._sheetname_getter = sheetname_getter
        self._columns = key_column_map

    def connect(self, *args, **kwargs):
        self._wb = self._workbook_loader()
        self._sht = self._wb[self._sheetname_getter(self._wb.sheetnames)]

    def close(self):
        self._sht = None
        self._wb.close()

    def insert(self, *args, sync=True, **kwargs):
        item = {self._columns[key]: self._sheet_format(key, value) for key, value in kwargs.get("item").items()}
        with self:
            self._sht.append(item)
            if sync and self._filename is not None:
                self.sync_to_file()

    def sync_to_file(self):
        self._wb.save(filename=self._filename)

    def update(self, *args, **kwargs):
        pass

    def _process_get_params(self, *args, **kwargs):
        """
        Returns pre_process: List[List[Any]] -> List[List[Any]],
                filter_predicate : List[Any] -> bool,
                post_process: List[List[Any]] -> List[List[Any]]
        """
        action, *params = args
        action = action.lower()
        supported = ['where', 'all']
        if action not in supported:
            raise NotImplementedError(
                f"{self.__class__.__name__} supports only {','.join(supported)} (e.g get('where', 'name' , '==', 'myname'))")
        if action == 'all':
            return lambda x: x, lambda r: True, lambda x: x

        key, comparison, value = params

        key = key.lower()
        assert key in self._columns.keys()
        col = col_idx(self._columns[key])
        value = self._sheet_format(key, value)

        comparisons = {'==': operator.eq, '!=': operator.ne, '>=': operator.ge, '<=': operator.le, '>': operator.gt,
                       '<': operator.lt}
        assert comparison in comparisons.keys()
        cmp_func = comparisons[comparison]

        pre_process = lambda x: x
        if 'order_by' in kwargs:
            sortby_col = col_idx(self._columns[kwargs['order_by']])
            pre_process = lambda rows: sorted(rows, key=lambda row: row[sortby_col], reverse=kwargs.get('reverse', False))

        filter_predicate = lambda row: cmp_func(row[col], value)
        post_process = lambda x: x
        if 'limit' in kwargs:
            post_process = lambda rows: islice(rows, 0, kwargs['limit'])
        return pre_process, filter_predicate, post_process

    def get(self, *args, **kwargs):
        """
        'where' is supported with '==', '!=', '>=', '<=', '>', '<'
        'all' is supported without additional args
        optional order_by, reverse and limit
        Example: get('where', 'age', '>=', 18, order_by='age', reverse=False, limit=5)  # youngest 5 above 18
        'age' must be a column found in 'columns' dict at __init__
        """
        pre_process, filter_predicate, post_process = self._process_get_params(*args, **kwargs)

        with self:
            cells_to_values = lambda row: list(map(lambda cell: cell.value, row))
            rows = map(cells_to_values, skip_first(self._sht.iter_rows()))   # skip first headers row
            pre_processed = pre_process(rows)
            filtered = list(filter(filter_predicate, pre_processed))
            post_processed = post_process(filtered)
            user_format_value = lambda row, key, col_letter: self._from_sheet_format(key, row[col_idx(col_letter)])
            res = [{key: user_format_value(row, key, col) for key, col in self._columns.items()} for row in post_processed]
            return res

    def _sheet_format(self, key, value):
        if type(value) is bool:
            assert key in self._boolean_keys, f'Key {key} is not in boolean_keys'
            return int(value)
        return value

    def _from_sheet_format(self, key, value):
        if key in self._boolean_keys:
            return bool(value)
        return value

    def __enter__(self):
        self.connect()
        return self

    def __exit__(self, exc_type, exc_value, tb):
        if exc_type is not None:
            # traceback.print_exception(exc_type, exc_value, tb)
            return False  # uncomment to pass exception through

        return True
